"""
Report service.

Builds a styled multi-sheet openpyxl workbook and returns it as a
BytesIO buffer.  The caller wraps that buffer in a StreamingResponse.

IMPORTANT NOTES:
  1. asyncio.gather() is NOT used — all repo calls share one AsyncSession
     which is not safe for concurrent access (SQLAlchemy raises
     InvalidRequestError if two coroutines touch the session simultaneously).
  2. All sheet builders receive plain Row objects from raw SQL queries
     (not ORM instances), so attributes are accessed directly by column name.
"""

from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.reports_repo import ReportRepo
from app.schemas.reports_schemas import ReportFilters

# ─────────────────────────────────────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────────────────────────────────────

_TEAL   = "1A7F6E"
_WHITE  = "FFFFFF"
_SLATE  = "F1F5F9"
_BORDER = "CBD5E1"

_HEADER_FONT  = Font(name="Calibri", bold=True, color=_WHITE, size=10)
_HEADER_FILL  = PatternFill("solid", fgColor=_TEAL)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALT_FILL     = PatternFill("solid", fgColor=_SLATE)
_THIN_SIDE    = Side(style="thin", color=_BORDER)
_THIN         = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)
_DATA_FONT    = Font(name="Calibri", size=10)
_DATA_ALIGN   = Alignment(vertical="center")
_TITLE_FONT   = Font(name="Calibri", bold=True, size=13, color=_TEAL)
_META_FONT    = Font(name="Calibri", size=9, color="64748B")


# ─────────────────────────────────────────────────────────────────────────────
# Worksheet helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fmt(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    return str(v)


def _write_header(ws: Worksheet, columns: List[str], row: int = 1) -> None:
    for col_idx, label in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.border    = _THIN
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[row].height = 22


def _write_row(ws: Worksheet, values: List[Any], row: int, alt: bool) -> None:
    fill = _ALT_FILL if alt else None
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=_fmt(val))
        cell.font      = _DATA_FONT
        cell.border    = _THIN
        cell.alignment = _DATA_ALIGN
        if fill:
            cell.fill = fill


def _auto_width(ws: Worksheet, min_w: int = 12, max_w: int = 40) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 2, min_w), max_w
        )


def _freeze(ws: Worksheet, row: int = 2) -> None:
    ws.freeze_panes = ws.cell(row=row, column=1)


def _sheet_meta(
    ws: Worksheet, title: str, subtitle: str, filters: ReportFilters
) -> int:
    """Write title block, return the row number at which the header row goes."""
    ws.merge_cells("A1:F1")
    ws["A1"].value     = title
    ws["A1"].font      = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:F2")
    parts = [f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"]
    if filters.date_from:
        parts.append(f"From: {filters.date_from}")
    if filters.date_to:
        parts.append(f"To: {filters.date_to}")
    if filters.patient_type:
        parts.append(f"Type: {filters.patient_type}")
    if filters.patient_status:
        parts.append(f"Status: {filters.patient_status}")
    ws["A2"].value = "  |  ".join(parts)
    ws["A2"].font  = _META_FONT
    ws.row_dimensions[2].height = 16

    return 3   # header goes on row 3, data starts on row 4


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders  (all receive plain Row objects from raw SQL)
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary(ws: Worksheet, stats: dict) -> None:
    ws["A1"].value = "HepVac — Export Summary"
    ws["A1"].font  = _TITLE_FONT
    ws.row_dimensions[1].height = 28
    ws["A2"].value = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font  = _META_FONT

    _write_header(ws, ["Sheet", "Records Exported"], row=4)
    for i, (sheet, count) in enumerate(stats.items(), start=5):
        _write_row(ws, [sheet, count], row=i, alt=(i % 2 == 0))

    _auto_width(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def _build_patients(ws: Worksheet, rows, filters: ReportFilters) -> int:
    data_row = _sheet_meta(ws, "Patient Register", "All registered patients", filters)
    cols = [
        "Patient ID", "Name", "Phone", "Sex", "Date of Birth", "Age",
        "Type", "Status", "Facility", "Registered On", "Accepts Messaging",
        "Gravida", "Para",
    ]
    _write_header(ws, cols, row=data_row)
    data_row += 1

    today = date.today()
    for i, r in enumerate(rows):
        age = ""
        if r.date_of_birth:
            dob = r.date_of_birth if isinstance(r.date_of_birth, date) else r.date_of_birth.date()
            age = (
                today.year - dob.year
                - ((today.month, today.day) < (dob.month, dob.day))
            )
        _write_row(ws, [
            str(r.id), r.name, r.phone, r.sex, r.date_of_birth, age,
            r.patient_type, r.patient_status, r.facility_name,
            r.created_at, r.accepts_messaging,
            r.gravida, r.para,
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


def _build_pregnancies(ws: Worksheet, rows, filters: ReportFilters) -> int:
    data_row = _sheet_meta(ws, "Pregnancy Register", "Per-episode obstetric records", filters)
    cols = [
        "Pregnancy ID", "Patient ID", "Patient Name", "Pregnancy #",
        "LMP Date", "Expected Delivery", "Actual Delivery",
        "Gestational Age (wks)", "Is Active", "Outcome",
        "Risk Factors", "Notes", "Created",
    ]
    _write_header(ws, cols, row=data_row)
    data_row += 1

    for i, r in enumerate(rows):
        _write_row(ws, [
            str(r.id), str(r.patient_id), r.patient_name, r.pregnancy_number,
            r.lmp_date, r.expected_delivery_date, r.actual_delivery_date,
            r.gestational_age_weeks,
            "Yes" if r.is_active else "No",
            r.outcome,
            r.risk_factors, r.notes, r.created_at,
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


def _build_children(ws: Worksheet, rows, filters: ReportFilters) -> int:
    data_row = _sheet_meta(ws, "Children Register", "Birth & monitoring records", filters)
    cols = [
        "Child ID", "Name", "Sex", "Date of Birth",
        "Mother Patient ID", "Mother Name", "Pregnancy ID",
        "6mo Checkup Date", "Checkup Completed",
        "Hep B Antibody Result", "Test Date", "Notes",
    ]
    _write_header(ws, cols, row=data_row)
    data_row += 1

    for i, r in enumerate(rows):
        _write_row(ws, [
            str(r.id), r.name, r.sex, r.date_of_birth,
            str(r.mother_patient_id), r.mother_name, str(r.pregnancy_id),
            r.six_month_checkup_date,
            "Yes" if r.six_month_checkup_completed else "No",
            r.hep_b_antibody_test_result,
            r.test_date, r.notes,
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


def _build_transactions(ws: Worksheet, rows, filters: ReportFilters) -> int:
    data_row = _sheet_meta(
        ws, "Vaccine Transactions", "Purchases & instalment payments", filters
    )
    cols = [
        "Purchase ID", "Patient ID", "Patient Name", "Phone",
        "Vaccine", "Batch #", "Total Doses",
        "Price/Dose (GHS)", "Total Package (GHS)", "Amount Paid (GHS)",
        "Balance (GHS)", "Payment Status",
        "Doses Administered", "Purchase Date", "Is Active",
    ]
    _write_header(ws, cols, row=data_row)
    data_row += 1

    for i, r in enumerate(rows):
        _write_row(ws, [
            str(r.id), str(r.patient_id), r.patient_name, r.patient_phone,
            r.vaccine_name, r.batch_number, r.total_doses,
            str(r.price_per_dose), str(r.total_package_price),
            str(r.amount_paid), str(r.balance),
            r.payment_status,
            r.doses_administered, r.purchase_date,
            "Yes" if r.is_active else "No",
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


def _build_vaccinations(ws: Worksheet, rows, filters: ReportFilters) -> int:
    data_row = _sheet_meta(ws, "Dose Administration Log", "Individual doses given", filters)
    cols = [
        "Vaccination ID", "Patient ID", "Patient Name",
        "Vaccine Name", "Dose #", "Dose Date", "Batch #",
        "Price Charged (GHS)", "Administered By", "Notes",
    ]
    _write_header(ws, cols, row=data_row)
    data_row += 1

    for i, r in enumerate(rows):
        _write_row(ws, [
            str(r.id), str(r.patient_id), r.patient_name,
            r.vaccine_name, r.dose_number, r.dose_date, r.batch_number,
            str(r.vaccine_price), r.administered_by, r.notes,
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


def _build_prescriptions(ws: Worksheet, rows, filters: ReportFilters) -> int:
    data_row = _sheet_meta(ws, "Prescriptions", "Medication prescriptions issued", filters)
    cols = [
        "Prescription ID", "Patient ID", "Patient Name",
        "Medication", "Dosage", "Frequency", "Duration (months)",
        "Prescription Date", "Start Date", "End Date",
        "Is Active", "Prescribed By", "Instructions",
    ]
    _write_header(ws, cols, row=data_row)
    data_row += 1

    for i, r in enumerate(rows):
        _write_row(ws, [
            str(r.id), str(r.patient_id), r.patient_name,
            r.medication_name, r.dosage, r.frequency, r.duration_months,
            r.prescription_date, r.start_date, r.end_date,
            "Yes" if r.is_active else "No",
            r.prescribed_by, r.instructions,
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


def _build_medication_schedules(ws: Worksheet, rows, filters: ReportFilters) -> int:
    data_row = _sheet_meta(
        ws, "Medication Schedules", "Monthly dispensing & follow-up records", filters
    )
    cols = [
        "Schedule ID", "Patient ID", "Patient Name",
        "Medication", "Scheduled Date", "Qty Purchased", "Months Supply",
        "Next Dose Due", "Completed", "Completed Date",
        "Lab Review Scheduled", "Lab Review Date", "Lab Review Done",
        "Notes",
    ]
    _write_header(ws, cols, row=data_row)
    data_row += 1

    for i, r in enumerate(rows):
        _write_row(ws, [
            str(r.id), str(r.patient_id), r.patient_name,
            r.medication_name, r.scheduled_date,
            r.quantity_purchased, r.months_supply,
            r.next_dose_due_date,
            "Yes" if r.is_completed else "No",
            r.completed_date,
            "Yes" if r.lab_review_scheduled else "No",
            r.lab_review_date,
            "Yes" if r.lab_review_completed else "No",
            r.notes,
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


def _build_diagnoses(ws: Worksheet, rows, filters: ReportFilters) -> int:
    data_row = _sheet_meta(ws, "Diagnoses", "Clinical diagnosis records", filters)
    cols = [
        "Diagnosis ID", "Patient ID", "Patient Name",
        "History", "Preliminary Diagnosis", "Actual Diagnosis",
        "Diagnosed On", "Diagnosed By",
    ]
    _write_header(ws, cols, row=data_row)
    data_row += 1

    for i, r in enumerate(rows):
        _write_row(ws, [
            str(r.id), str(r.patient_id), r.patient_name,
            r.history, r.preliminary_diagnosis, r.actual_diagnosis,
            r.diagnosed_on, r.diagnosed_by,
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


def _build_reminders(ws: Worksheet, rows, filters: ReportFilters) -> int:
    data_row = _sheet_meta(ws, "Reminders", "Automated reminder records", filters)
    cols = [
        "Reminder ID", "Patient ID", "Patient Name", "Phone",
        "Type", "Scheduled Date", "Status", "Sent At", "Message",
    ]
    _write_header(ws, cols, row=data_row)
    data_row += 1

    for i, r in enumerate(rows):
        _write_row(ws, [
            str(r.id), str(r.patient_id), r.patient_name, r.patient_phone,
            r.reminder_type, r.scheduled_date, r.status, r.sent_at, r.message,
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


def _build_vaccine_stock(ws: Worksheet, rows) -> int:
    ws["A1"].value = "Vaccine Stock"
    ws["A1"].font  = _TITLE_FONT
    ws.row_dimensions[1].height = 24
    ws["A2"].value = f"Snapshot: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font  = _META_FONT

    cols = [
        "Vaccine ID", "Name", "Price/Dose (GHS)", "Total Stock",
        "Reserved Qty", "Available Qty", "Batch #", "Published", "Created",
    ]
    _write_header(ws, cols, row=3)
    data_row = 4

    for i, r in enumerate(rows):
        low = (r.available_quantity or 0) <= 10
        _write_row(ws, [
            str(r.id), r.vaccine_name, str(r.price_per_dose),
            r.quantity, r.reserved_quantity, r.available_quantity,
            r.batch_number,
            "Yes" if r.is_published else "No",
            "⚠ Low Stock" if low else "OK",
            r.created_at,
        ], row=data_row, alt=(i % 2 == 1))
        data_row += 1

    _auto_width(ws)
    _freeze(ws, row=4)
    return len(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Main service
# ─────────────────────────────────────────────────────────────────────────────

class ReportService:
    def __init__(self, db: AsyncSession) -> None:
        self._repo = ReportRepo(db)

    async def build_workbook(self, filters: ReportFilters) -> BytesIO:
        """
        Fetch data sequentially (one await at a time — AsyncSession is not
        safe for concurrent operations), build the workbook, return as BytesIO.
        """
        results: dict[str, list] = {}

        # ── Sequential fetches ────────────────────────────────────────────
        # DO NOT convert to asyncio.gather() — shared session, single connection.
        if filters.include_patients:
            results["patients"] = await self._repo.get_patients(filters)
        if filters.include_pregnancies:
            results["pregnancies"] = await self._repo.get_pregnancies(filters)
        if filters.include_children:
            results["children"] = await self._repo.get_children(filters)
        if filters.include_transactions:
            results["purchases"] = await self._repo.get_purchases(filters)
        if filters.include_vaccinations:
            results["vaccinations"] = await self._repo.get_vaccinations(filters)
        if filters.include_prescriptions:
            results["prescriptions"] = await self._repo.get_prescriptions(filters)
        if filters.include_medications:
            results["schedules"] = await self._repo.get_medication_schedules(filters)
        if filters.include_diagnoses:
            results["diagnoses"] = await self._repo.get_diagnoses(filters)
        if filters.include_reminders:
            results["reminders"] = await self._repo.get_reminders(filters)
        if filters.include_stock:
            results["vaccines"] = await self._repo.get_vaccines()

        # ── Build workbook ────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove the default empty sheet

        stats: dict[str, int] = {}

        if filters.include_patients:
            ws = wb.create_sheet("Patients")
            stats["Patients"] = _build_patients(ws, results["patients"], filters)
        if filters.include_pregnancies:
            ws = wb.create_sheet("Pregnancies")
            stats["Pregnancies"] = _build_pregnancies(ws, results["pregnancies"], filters)
        if filters.include_children:
            ws = wb.create_sheet("Children")
            stats["Children"] = _build_children(ws, results["children"], filters)
        if filters.include_transactions:
            ws = wb.create_sheet("Transactions")
            stats["Transactions"] = _build_transactions(ws, results["purchases"], filters)
        if filters.include_vaccinations:
            ws = wb.create_sheet("Vaccinations")
            stats["Vaccinations"] = _build_vaccinations(ws, results["vaccinations"], filters)
        if filters.include_prescriptions:
            ws = wb.create_sheet("Prescriptions")
            stats["Prescriptions"] = _build_prescriptions(ws, results["prescriptions"], filters)
        if filters.include_medications:
            ws = wb.create_sheet("Medication Schedules")
            stats["Medication Schedules"] = _build_medication_schedules(
                ws, results["schedules"], filters
            )
        if filters.include_diagnoses:
            ws = wb.create_sheet("Diagnoses")
            stats["Diagnoses"] = _build_diagnoses(ws, results["diagnoses"], filters)
        if filters.include_reminders:
            ws = wb.create_sheet("Reminders")
            stats["Reminders"] = _build_reminders(ws, results["reminders"], filters)
        if filters.include_stock:
            ws = wb.create_sheet("Vaccine Stock")
            stats["Vaccine Stock"] = _build_vaccine_stock(ws, results["vaccines"])

        # Summary sheet goes first
        summary_ws = wb.create_sheet("Summary", 0)
        _build_summary(summary_ws, stats)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
